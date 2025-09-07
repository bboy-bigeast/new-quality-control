"""通用工具函数模块 - 用于提取重复的统计计算逻辑"""

import csv
import io
from django.http import HttpResponse
from django.db.models import Avg, StdDev
import numpy as np
from scipy import stats
from products.models import ProductStandard
import openpyxl
from openpyxl.utils import get_column_letter

def calculate_statistics(data_values):
    """计算基本统计信息"""
    valid_data = [x for x in data_values if x is not None]
    if not valid_data:
        return {
            'average': 0,
            'std_dev': 0,
            'std_dev_lines': {
                'plus_1sigma': 0, 'minus_1sigma': 0,
                'plus_2sigma': 0, 'minus_2sigma': 0,
                'plus_3sigma': 0, 'minus_3sigma': 0,
                'plus_4sigma': 0, 'minus_4sigma': 0,
                'plus_5sigma': 0, 'minus_5sigma': 0,
            }
        }
    
    avg = sum(valid_data) / len(valid_data)
    std_dev = (sum((x - avg) ** 2 for x in valid_data) / len(valid_data)) ** 0.5
    
    return {
        'average': avg,
        'std_dev': std_dev,
        'std_dev_lines': {
            'plus_1sigma': avg + std_dev,
            'minus_1sigma': avg - std_dev,
            'plus_2sigma': avg + 2 * std_dev,
            'minus_2sigma': avg - 2 * std_dev,
            'plus_3sigma': avg + 3 * std_dev,
            'minus_3sigma': avg - 3 * std_dev,
            'plus_4sigma': avg + 4 * std_dev,
            'minus_4sigma': avg - 4 * std_dev,
            'plus_5sigma': avg + 5 * std_dev,
            'minus_5sigma': avg - 5 * std_dev,
        }
    }

def get_product_field_value(product, test_item, product_type='dryfilm'):
    """根据测试项目获取产品字段值"""
    field_mapping = {
        'dryfilm': {
            'solid_content': 'solid_content',
            'viscosity': 'viscosity',
            'acid_value': 'acid_value',
            'moisture': 'moisture',
            'residual_monomer': 'residual_monomer',
            'weight_avg_molecular_weight': 'weight_avg_molecular_weight',
            'pdi': 'pdi',
            'color': 'color',
        },
        'adhesive': {
            'solid_content': 'solid_content',
            'viscosity': 'viscosity',
            'acid_value': 'acid_value',
            'moisture': 'moisture',
            'residual_monomer': 'residual_monomer',
            'weight_avg_molecular_weight': 'weight_avg_molecular_weight',
            'pdi': 'pdi',
            'color': 'color',
            'initial_tack': 'initial_tack',
            'peel_strength': 'peel_strength',
            'high_temperature_holding': 'high_temperature_holding',
            'room_temperature_holding': 'room_temperature_holding',
            'constant_load_peel': 'constant_load_peel',
        }
    }
    
    field_name = field_mapping[product_type].get(test_item)
    return getattr(product, field_name, None) if field_name else None

def calculate_moving_range_data(data_values):
    """计算移动极差数据"""
    moving_ranges = []
    previous_value = None
    
    for current_value in data_values:
        if previous_value is not None and current_value is not None:
            moving_range = abs(current_value - previous_value)
            moving_ranges.append(moving_range)
        else:
            moving_ranges.append(None)
        previous_value = current_value
    
    # 计算移动极差统计
    valid_moving_ranges = [x for x in moving_ranges if x is not None]
    if valid_moving_ranges:
        mr_avg = sum(valid_moving_ranges) / len(valid_moving_ranges)
        ucl_mr = 3.267 * mr_avg  # 移动极差图上控制限
    else:
        mr_avg = 0
        ucl_mr = 0
    
    return {
        'moving_ranges': moving_ranges,
        'statistics': {
            'mr_average': mr_avg,
            'ucl_mr': ucl_mr,
        }
    }

def calculate_capability_analysis(data_values, product_code=None, test_item=None):
    """计算能力分析数据"""
    if not data_values or len(data_values) <= 1:
        return {
            'histogram': {'values': [], 'bins': []},
            'normal_distribution': {'x': [], 'y': []},
            'statistics': {
                'mean': 0, 'std_dev': 0, 'usl': None, 'lsl': None,
                'target': 0, 'cp': None, 'cpk': None, 'sample_size': 0
            }
        }
    
    data_array = np.array(data_values)
    mean = float(np.mean(data_array))
    std_dev = float(np.std(data_array, ddof=1))  # 样本标准差
    
    # 获取产品标准
    usl, lsl, target = get_product_standard(product_code, test_item, mean)
    
    # 计算过程能力指数
    cp, cpk = calculate_process_capability(mean, std_dev, usl, lsl)
    
    # 生成正态分布曲线数据
    x_min = float(min(data_values) - 3 * std_dev)
    x_max = float(max(data_values) + 3 * std_dev)
    x = np.linspace(x_min, x_max, 100)
    y = stats.norm.pdf(x, mean, std_dev)
    
    # 计算直方图数据
    hist, bin_edges = np.histogram(data_values, bins=min(10, len(data_values)), density=True)
    
    return {
        'histogram': {
            'values': [float(v) for v in hist.tolist()],
            'bins': [float(b) for b in bin_edges.tolist()]
        },
        'normal_distribution': {
            'x': [float(v) for v in x.tolist()],
            'y': [float(v) for v in y.tolist()]
        },
        'statistics': {
            'mean': mean,
            'std_dev': std_dev,
            'usl': usl,
            'lsl': lsl,
            'target': target,
            'cp': cp,
            'cpk': cpk,
            'sample_size': len(data_values)
        }
    }

def get_product_standard(product_code, test_item, default_target):
    """获取产品标准限制"""
    if not product_code or not test_item:
        return None, None, default_target
    
    try:
        standard = ProductStandard.objects.get(
            product_code=product_code,
            test_item=test_item,
            standard_type='internal_control'
        )
        usl = float(standard.upper_limit) if standard.upper_limit else None
        lsl = float(standard.lower_limit) if standard.lower_limit else None
        target = float(standard.target_value) if standard.target_value else default_target
        return usl, lsl, target
    except ProductStandard.DoesNotExist:
        return None, None, default_target

def calculate_process_capability(mean, std_dev, usl, lsl):
    """计算过程能力指数"""
    if usl is None or lsl is None or std_dev <= 0:
        return None, None
    
    cp = float((usl - lsl) / (6 * std_dev))  # 过程能力指数
    cpk = float(min((usl - mean) / (3 * std_dev), (mean - lsl) / (3 * std_dev)))  # 实际过程能力指数
    return cp, cpk

def get_batch_date(product, date_field='test_date'):
    """获取批次日期"""
    if hasattr(product, 'batch_number') and len(getattr(product, 'batch_number', '')) >= 8:
        return product.batch_number[:8]
    else:
        date_value = getattr(product, date_field)
        return date_value.strftime('%Y%m%d') if date_value else '00000000'

def export_to_csv(model_name, queryset, fields):
    """导出数据到CSV格式"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{model_name}_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow(fields)
    
    for obj in queryset:
        row = []
        for field in fields:
            value = getattr(obj, field, '')
            if hasattr(value, '__call__'):
                value = value()
            elif hasattr(value, 'strftime'):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            row.append(str(value))
        writer.writerow(row)
    
    return response

def export_to_excel(model_name, queryset, fields):
    """导出数据到Excel格式"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model_name
    
    # 写入表头
    for col_num, field in enumerate(fields, 1):
        ws.cell(row=1, column=col_num, value=field)
    
    # 写入数据
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field, '')
            if hasattr(value, '__call__'):
                value = value()
            elif hasattr(value, 'strftime'):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_num, column=col_num, value=str(value))
    
    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{model_name}_export.xlsx"'
    
    buffer = io.BytesIO()
    wb.save(buffer)
    response.write(buffer.getvalue())
    
    return response

def export_data(request, queryset, model_name, fields, format_type='csv'):
    """通用数据导出函数"""
    if format_type == 'csv':
        return export_to_csv(model_name, queryset, fields)
    elif format_type == 'excel':
        return export_to_excel(model_name, queryset, fields)
    else:
        return HttpResponse("不支持的导出格式", status=400)
